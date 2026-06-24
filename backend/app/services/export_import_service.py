"""
Data Export/Import Service
Enterprise-grade data portability and migration
"""

import csv
import json
import logging
import hashlib
import zipfile
import pandas as pd
from io import BytesIO, StringIO
from typing import Dict, Any, List, Optional, BinaryIO
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_

from ..models import Project, Transcript, User, Org, _uid
from ..models_phase2 import AnalysisGrid, GridCell, AnalysisTheme, Evidence
from ..models_phase3 import AnalysisInsight
from ..models_phase4 import Comment
from ..models_phase5 import Clip, Reel
from ..models_enterprise import DataExportJob, DataImportJob
from ..config import settings
from ..tasks import celery_app

logger = logging.getLogger(__name__)


class DataExportService:
    """Service for exporting data in various formats"""

    SUPPORTED_FORMATS = ["json", "csv", "xlsx", "pdf", "spss", "nvivo", "maxqda"]

    @classmethod
    def create_export_job(
        cls,
        db: Session,
        user: User,
        export_type: str,
        format: str,
        project_ids: Optional[List[str]] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        include_transcripts: bool = True,
        include_analysis: bool = True,
        include_media: bool = False,
        include_deleted: bool = False,
        anonymize_pii: bool = False,
        language: str = "en",
        timezone: str = "UTC"
    ) -> DataExportJob:
        """Create a new export job"""

        if format not in cls.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {format}")

        job = DataExportJob(
            id=_uid(),
            org_id=user.org_id,
            user_id=user.id,
            export_type=export_type,
            format=format,
            project_ids=project_ids,
            date_from=date_from,
            date_to=date_to,
            include_transcripts=include_transcripts,
            include_analysis=include_analysis,
            include_media=include_media,
            include_deleted=include_deleted,
            anonymize_pii=anonymize_pii,
            language=language,
            timezone=timezone,
            status="pending"
        )

        db.add(job)
        db.commit()
        db.refresh(job)

        # Queue async task
        task = celery_app.send_task(
            "app.tasks.process_export",
            args=[job.id]
        )
        job.celery_task_id = task.id
        db.commit()

        return job

    @classmethod
    def export_to_json(
        cls,
        db: Session,
        job: DataExportJob
    ) -> tuple[str, int]:
        """Export data to JSON format"""

        data = {
            "export_info": {
                "export_id": job.id,
                "export_date": datetime.utcnow().isoformat(),
                "export_type": job.export_type,
                "org_id": job.org_id,
                "user_id": job.user_id
            },
            "projects": [],
            "metadata": {}
        }

        # Build project query
        query = db.query(Project).filter(Project.org_id == job.org_id)

        if job.project_ids:
            query = query.filter(Project.id.in_(job.project_ids))

        if job.date_from:
            query = query.filter(Project.created_at >= job.date_from)

        if job.date_to:
            query = query.filter(Project.created_at <= job.date_to)

        projects = query.all()
        records_count = 0

        for project in projects:
            project_data = {
                "id": project.id,
                "name": project.name,
                "description": project.description,
                "created_at": project.created_at.isoformat(),
                "updated_at": project.updated_at.isoformat()
            }

            # Include transcripts
            if job.include_transcripts:
                transcripts = db.query(Transcript).filter(
                    Transcript.project_id == project.id
                ).all()

                project_data["transcripts"] = []
                for transcript in transcripts:
                    transcript_data = {
                        "id": transcript.id,
                        "name": transcript.name,
                        "content": transcript.content if not job.anonymize_pii else cls._anonymize_text(transcript.content),
                        "speaker": transcript.speaker if not job.anonymize_pii else "SPEAKER",
                        "duration": transcript.duration,
                        "created_at": transcript.created_at.isoformat()
                    }
                    project_data["transcripts"].append(transcript_data)
                    records_count += 1

            # Include analysis
            if job.include_analysis:
                # Grids
                grids = db.query(AnalysisGrid).filter(
                    AnalysisGrid.project_id == project.id
                ).all()

                project_data["analysis_grids"] = []
                for grid in grids:
                    grid_data = {
                        "id": grid.id,
                        "name": grid.name,
                        "markets": grid.markets,
                        "created_at": grid.created_at.isoformat(),
                        "cells": []
                    }

                    cells = db.query(GridCell).filter(
                        GridCell.grid_id == grid.id
                    ).all()

                    for cell in cells:
                        cell_data = {
                            "id": cell.id,
                            "row_label": cell.row_label,
                            "column_label": cell.column_label,
                            "content": cell.content if not job.anonymize_pii else cls._anonymize_text(cell.content)
                        }
                        grid_data["cells"].append(cell_data)

                    project_data["analysis_grids"].append(grid_data)
                    records_count += len(cells)

                # Themes
                themes = db.query(AnalysisTheme).filter(
                    AnalysisTheme.project_id == project.id
                ).all()

                project_data["themes"] = []
                for theme in themes:
                    theme_data = {
                        "id": theme.id,
                        "name": theme.name,
                        "description": theme.description,
                        "prevalence": theme.prevalence,
                        "created_at": theme.created_at.isoformat()
                    }
                    project_data["themes"].append(theme_data)
                    records_count += 1

                # Insights
                insights = db.query(AnalysisInsight).filter(
                    AnalysisInsight.project_id == project.id
                ).all()

                project_data["insights"] = []
                for insight in insights:
                    insight_data = {
                        "id": insight.id,
                        "title": insight.title,
                        "description": insight.description,
                        "category": insight.category,
                        "confidence": insight.confidence,
                        "created_at": insight.created_at.isoformat()
                    }
                    project_data["insights"].append(insight_data)
                    records_count += 1

            data["projects"].append(project_data)

        # Add metadata
        data["metadata"]["total_projects"] = len(projects)
        data["metadata"]["total_records"] = records_count

        # Save to file
        export_path = Path(settings.storage_path) / "exports" / f"{job.id}.json"
        export_path.parent.mkdir(parents=True, exist_ok=True)

        with open(export_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return str(export_path), records_count

    @classmethod
    def export_to_csv(
        cls,
        db: Session,
        job: DataExportJob
    ) -> tuple[str, int]:
        """Export data to CSV format (multiple files in ZIP)"""

        export_dir = Path(settings.storage_path) / "exports" / job.id
        export_dir.mkdir(parents=True, exist_ok=True)

        records_count = 0

        # Export projects
        projects_query = db.query(Project).filter(Project.org_id == job.org_id)
        if job.project_ids:
            projects_query = projects_query.filter(Project.id.in_(job.project_ids))

        projects = projects_query.all()
        projects_data = []

        for project in projects:
            projects_data.append({
                "id": project.id,
                "name": project.name,
                "description": project.description,
                "created_at": project.created_at,
                "updated_at": project.updated_at
            })

        if projects_data:
            df = pd.DataFrame(projects_data)
            df.to_csv(export_dir / "projects.csv", index=False, encoding="utf-8-sig")
            records_count += len(projects_data)

        # Export transcripts
        if job.include_transcripts:
            transcripts_data = []
            for project in projects:
                transcripts = db.query(Transcript).filter(
                    Transcript.project_id == project.id
                ).all()

                for transcript in transcripts:
                    transcripts_data.append({
                        "project_id": project.id,
                        "project_name": project.name,
                        "transcript_id": transcript.id,
                        "transcript_name": transcript.name,
                        "speaker": transcript.speaker if not job.anonymize_pii else "SPEAKER",
                        "content": transcript.content if not job.anonymize_pii else cls._anonymize_text(transcript.content),
                        "duration": transcript.duration,
                        "created_at": transcript.created_at
                    })

            if transcripts_data:
                df = pd.DataFrame(transcripts_data)
                df.to_csv(export_dir / "transcripts.csv", index=False, encoding="utf-8-sig")
                records_count += len(transcripts_data)

        # Export analysis grids
        if job.include_analysis:
            grids_data = []
            cells_data = []

            for project in projects:
                grids = db.query(AnalysisGrid).filter(
                    AnalysisGrid.project_id == project.id
                ).all()

                for grid in grids:
                    grids_data.append({
                        "project_id": project.id,
                        "project_name": project.name,
                        "grid_id": grid.id,
                        "grid_name": grid.name,
                        "markets": ", ".join(grid.markets) if grid.markets else "",
                        "created_at": grid.created_at
                    })

                    cells = db.query(GridCell).filter(
                        GridCell.grid_id == grid.id
                    ).all()

                    for cell in cells:
                        cells_data.append({
                            "grid_id": grid.id,
                            "grid_name": grid.name,
                            "row_label": cell.row_label,
                            "column_label": cell.column_label,
                            "content": cell.content if not job.anonymize_pii else cls._anonymize_text(cell.content),
                            "created_at": cell.created_at
                        })

            if grids_data:
                df = pd.DataFrame(grids_data)
                df.to_csv(export_dir / "analysis_grids.csv", index=False, encoding="utf-8-sig")
                records_count += len(grids_data)

            if cells_data:
                df = pd.DataFrame(cells_data)
                df.to_csv(export_dir / "grid_cells.csv", index=False, encoding="utf-8-sig")
                records_count += len(cells_data)

        # Create ZIP archive
        zip_path = Path(settings.storage_path) / "exports" / f"{job.id}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for file_path in export_dir.glob("*.csv"):
                zipf.write(file_path, file_path.name)

        # Clean up temporary files
        for file_path in export_dir.glob("*.csv"):
            file_path.unlink()
        export_dir.rmdir()

        return str(zip_path), records_count

    @classmethod
    def export_to_xlsx(
        cls,
        db: Session,
        job: DataExportJob
    ) -> tuple[str, int]:
        """Export data to Excel format with multiple sheets"""

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        records_count = 0

        # Export projects
        projects_query = db.query(Project).filter(Project.org_id == job.org_id)
        if job.project_ids:
            projects_query = projects_query.filter(Project.id.in_(job.project_ids))

        projects = projects_query.all()
        projects_data = []

        for project in projects:
            projects_data.append({
                "ID": project.id,
                "Name": project.name,
                "Description": project.description,
                "Created": project.created_at,
                "Updated": project.updated_at
            })

        if projects_data:
            ws = wb.create_sheet("Projects")
            df = pd.DataFrame(projects_data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            records_count += len(projects_data)

        # Export transcripts
        if job.include_transcripts:
            transcripts_data = []
            for project in projects:
                transcripts = db.query(Transcript).filter(
                    Transcript.project_id == project.id
                ).all()

                for transcript in transcripts:
                    transcripts_data.append({
                        "Project": project.name,
                        "Transcript": transcript.name,
                        "Speaker": transcript.speaker if not job.anonymize_pii else "SPEAKER",
                        "Content": transcript.content[:1000] if not job.anonymize_pii else cls._anonymize_text(transcript.content[:1000]),
                        "Duration": transcript.duration,
                        "Created": transcript.created_at
                    })

            if transcripts_data:
                ws = wb.create_sheet("Transcripts")
                df = pd.DataFrame(transcripts_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                records_count += len(transcripts_data)

        # Export themes
        if job.include_analysis:
            themes_data = []
            for project in projects:
                themes = db.query(AnalysisTheme).filter(
                    AnalysisTheme.project_id == project.id
                ).all()

                for theme in themes:
                    themes_data.append({
                        "Project": project.name,
                        "Theme": theme.name,
                        "Description": theme.description,
                        "Prevalence": theme.prevalence,
                        "Created": theme.created_at
                    })

            if themes_data:
                ws = wb.create_sheet("Themes")
                df = pd.DataFrame(themes_data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                records_count += len(themes_data)

        # Save workbook
        export_path = Path(settings.storage_path) / "exports" / f"{job.id}.xlsx"
        export_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(export_path)

        return str(export_path), records_count

    @classmethod
    def _anonymize_text(cls, text: str) -> str:
        """Anonymize PII in text"""
        import re

        # Replace email addresses
        text = re.sub(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', '[EMAIL]', text)

        # Replace phone numbers (various formats)
        text = re.sub(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b', '[PHONE]', text)
        text = re.sub(r'\b\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,4}\b', '[PHONE]', text)

        # Replace common ID numbers (SSN, passport, etc.)
        text = re.sub(r'\b\d{3}-\d{2}-\d{4}\b', '[ID]', text)  # SSN format
        text = re.sub(r'\b[A-Z]\d{7}\b', '[ID]', text)  # Passport format

        # Replace names (simple heuristic - words after Mr./Ms./Dr.)
        text = re.sub(r'\b(Mr\.|Ms\.|Mrs\.|Dr\.)\s+\w+\s+\w+', '[NAME]', text)

        return text


class DataImportService:
    """Service for importing data from various formats"""

    SUPPORTED_FORMATS = ["json", "csv", "xlsx", "nvivo", "maxqda", "atlas"]

    @classmethod
    def create_import_job(
        cls,
        db: Session,
        user: User,
        import_type: str,
        source_format: str,
        file_path: str,
        field_mappings: Optional[Dict[str, str]] = None,
        default_values: Optional[Dict[str, Any]] = None,
        validate_only: bool = False,
        skip_duplicates: bool = True,
        merge_strategy: str = "skip"
    ) -> DataImportJob:
        """Create a new import job"""

        if source_format not in cls.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {source_format}")

        # Calculate file hash
        file_hash = cls._calculate_file_hash(file_path)

        # Get file size
        file_size_mb = Path(file_path).stat().st_size / (1024 * 1024)

        job = DataImportJob(
            id=_uid(),
            org_id=user.org_id,
            user_id=user.id,
            import_type=import_type,
            source_format=source_format,
            file_name=Path(file_path).name,
            file_size_mb=file_size_mb,
            file_hash=file_hash,
            field_mappings=field_mappings or {},
            default_values=default_values or {},
            validate_only=validate_only,
            skip_duplicates=skip_duplicates,
            merge_strategy=merge_strategy,
            status="pending"
        )

        db.add(job)
        db.commit()
        db.refresh(job)

        # Queue async task
        task = celery_app.send_task(
            "app.tasks.process_import",
            args=[job.id, file_path]
        )
        job.celery_task_id = task.id
        db.commit()

        return job

    @classmethod
    def import_from_json(
        cls,
        db: Session,
        job: DataImportJob,
        file_path: str
    ) -> Dict[str, Any]:
        """Import data from JSON format"""

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        results = {
            "records_processed": 0,
            "records_imported": 0,
            "records_skipped": 0,
            "records_failed": 0,
            "validation_errors": [],
            "created_projects": [],
            "created_transcripts": []
        }

        # Validate structure
        if "projects" not in data:
            results["validation_errors"].append("Missing 'projects' key in JSON")
            return results

        for project_data in data["projects"]:
            results["records_processed"] += 1

            # Check for duplicates
            if job.skip_duplicates and "id" in project_data:
                existing = db.query(Project).filter(
                    Project.id == project_data["id"],
                    Project.org_id == job.org_id
                ).first()

                if existing:
                    if job.merge_strategy == "skip":
                        results["records_skipped"] += 1
                        continue
                    elif job.merge_strategy == "replace":
                        # Update existing project
                        existing.name = project_data.get("name", existing.name)
                        existing.description = project_data.get("description", existing.description)
                        existing.updated_at = datetime.utcnow()

            # Create new project
            if not job.validate_only:
                project = Project(
                    id=_uid(),
                    org_id=job.org_id,
                    name=project_data.get("name", "Imported Project"),
                    description=project_data.get("description", ""),
                    created_by=job.user_id
                )
                db.add(project)
                db.flush()

                results["created_projects"].append(project.id)

                # Import transcripts
                if "transcripts" in project_data:
                    for transcript_data in project_data["transcripts"]:
                        transcript = Transcript(
                            id=_uid(),
                            org_id=job.org_id,
                            project_id=project.id,
                            name=transcript_data.get("name", "Imported Transcript"),
                            content=transcript_data.get("content", ""),
                            speaker=transcript_data.get("speaker", "Unknown"),
                            duration=transcript_data.get("duration", 0)
                        )
                        db.add(transcript)
                        results["created_transcripts"].append(transcript.id)

                results["records_imported"] += 1

        if not job.validate_only:
            db.commit()

        return results

    @classmethod
    def import_from_csv(
        cls,
        db: Session,
        job: DataImportJob,
        file_path: str
    ) -> Dict[str, Any]:
        """Import data from CSV format"""

        results = {
            "records_processed": 0,
            "records_imported": 0,
            "records_skipped": 0,
            "records_failed": 0,
            "validation_errors": [],
            "created_projects": [],
            "created_transcripts": []
        }

        try:
            df = pd.read_csv(file_path, encoding="utf-8-sig")
        except Exception as e:
            results["validation_errors"].append(f"Failed to read CSV: {str(e)}")
            return results

        # Apply field mappings
        if job.field_mappings:
            df.rename(columns=job.field_mappings, inplace=True)

        # Process based on import type
        if job.import_type == "transcripts":
            for _, row in df.iterrows():
                results["records_processed"] += 1

                # Get or create project
                project_name = row.get("project_name", job.default_values.get("project_name", "Imported"))
                project = db.query(Project).filter(
                    Project.name == project_name,
                    Project.org_id == job.org_id
                ).first()

                if not project and not job.validate_only:
                    project = Project(
                        id=_uid(),
                        org_id=job.org_id,
                        name=project_name,
                        created_by=job.user_id
                    )
                    db.add(project)
                    db.flush()
                    results["created_projects"].append(project.id)

                # Create transcript
                if not job.validate_only and project:
                    transcript = Transcript(
                        id=_uid(),
                        org_id=job.org_id,
                        project_id=project.id,
                        name=row.get("transcript_name", "Imported Transcript"),
                        content=row.get("content", ""),
                        speaker=row.get("speaker", "Unknown")
                    )
                    db.add(transcript)
                    results["created_transcripts"].append(transcript.id)
                    results["records_imported"] += 1

        if not job.validate_only:
            db.commit()

        return results

    @classmethod
    def _calculate_file_hash(cls, file_path: str) -> str:
        """Calculate SHA256 hash of file"""
        sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256.update(chunk)
        return sha256.hexdigest()

    @classmethod
    def validate_import_file(
        cls,
        file_path: str,
        source_format: str
    ) -> tuple[bool, List[str]]:
        """Validate import file before processing"""

        errors = []

        # Check file exists
        if not Path(file_path).exists():
            errors.append("File does not exist")
            return False, errors

        # Check file size
        file_size_mb = Path(file_path).stat().st_size / (1024 * 1024)
        if file_size_mb > 100:  # 100MB limit
            errors.append(f"File too large: {file_size_mb:.2f}MB (max 100MB)")

        # Validate format-specific requirements
        if source_format == "json":
            try:
                with open(file_path, "r") as f:
                    data = json.load(f)
                    if not isinstance(data, dict):
                        errors.append("JSON must be an object")
            except json.JSONDecodeError as e:
                errors.append(f"Invalid JSON: {str(e)}")

        elif source_format == "csv":
            try:
                df = pd.read_csv(file_path, nrows=5)
                if df.empty:
                    errors.append("CSV file is empty")
            except Exception as e:
                errors.append(f"Invalid CSV: {str(e)}")

        elif source_format == "xlsx":
            try:
                df = pd.read_excel(file_path, nrows=5)
                if df.empty:
                    errors.append("Excel file is empty")
            except Exception as e:
                errors.append(f"Invalid Excel: {str(e)}")

        return len(errors) == 0, errors